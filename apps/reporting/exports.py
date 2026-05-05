import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from apps.interventions.models import Intervention
from django.utils import timezone


def export_excel(interventions):
    wb = Workbook()
    ws = wb.active
    ws.title = "Interventions OMEGA"

    # En-tête
    headers = ['#', 'Technicien', 'Prestation', 'Ville', 'Date', 'Statut', 'CA (€)', 'Gain Tech (€)', 'Marge (€)']
    header_fill = PatternFill("solid", fgColor="1e2130")
    header_font = Font(bold=True, color="4f8ef7")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Données
    for row, i in enumerate(interventions, 2):
        ws.cell(row=row, column=1, value=i.id)
        ws.cell(row=row, column=2, value=i.technicien.get_full_name())
        ws.cell(row=row, column=3, value=i.bareme.nom)
        ws.cell(row=row, column=4, value=i.ville)
        ws.cell(row=row, column=5, value=i.date_intervention.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=6, value=i.statut)
        ws.cell(row=row, column=7, value=float(i.chiffre_affaires))
        ws.cell(row=row, column=8, value=float(i.gain_technicien))
        ws.cell(row=row, column=9, value=float(i.marge))

    # Ligne totaux
    last = len(interventions) + 2
    ws.cell(row=last, column=6, value='TOTAL').font = Font(bold=True)
    ws.cell(row=last, column=7, value=sum(float(i.chiffre_affaires) for i in interventions)).font = Font(bold=True)
    ws.cell(row=last, column=8, value=sum(float(i.gain_technicien) for i in interventions)).font = Font(bold=True)
    ws.cell(row=last, column=9, value=sum(float(i.marge) for i in interventions)).font = Font(bold=True)

    # Largeur colonnes
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_pdf(interventions, periode):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Titre
    title_style = styles['Title']
    elements.append(Paragraph(f"Rapport OMEGA — {periode}", title_style))
    elements.append(Spacer(1, 0.5*cm))

    # Résumé
    total_ca = sum(float(i.chiffre_affaires) for i in interventions)
    total_cout = sum(float(i.gain_technicien) for i in interventions)
    total_marge = sum(float(i.marge) for i in interventions)
    nb = len(interventions)

    summary_data = [
        ['Interventions', 'Chiffre d\'affaires', 'Coûts techniciens', 'Marge brute'],
        [str(nb), f"{total_ca:.2f} €", f"{total_cout:.2f} €", f"{total_marge:.2f} €"]
    ]

    summary_table = Table(summary_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e2130')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#4f8ef7')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f0f4ff')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9ff')]),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.8*cm))

    # Tableau détail
    elements.append(Paragraph("Détail des interventions", styles['Heading2']))
    elements.append(Spacer(1, 0.3*cm))

    data = [['#', 'Technicien', 'Prestation', 'Ville', 'Date', 'CA (€)', 'Marge (€)']]
    for i in interventions:
        data.append([
            str(i.id),
            i.technicien.get_full_name(),
            i.bareme.nom,
            i.ville,
            i.date_intervention.strftime('%d/%m/%Y'),
            f"{float(i.chiffre_affaires):.2f}",
            f"{float(i.marge):.2f}",
        ])

    # Ligne total
    data.append(['', '', '', '', 'TOTAL', f"{total_ca:.2f}", f"{total_marge:.2f}"])

    table = Table(data, colWidths=[1*cm, 3.5*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e2130')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#4f8ef7')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9ff')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f0fe')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer